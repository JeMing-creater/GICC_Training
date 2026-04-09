from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Sequence

import pandas as pd
import torch
from monai.transforms import Activations, AsDiscrete
from openpyxl import load_workbook


# -------------------------
# Robust local imports
# -------------------------
try:
    from src.loader import get_loaders  # type: ignore
    from src.utils import load_cfg, load_weights, ensure_dir  # type: ignore
except Exception:
    from loader import get_loaders  # type: ignore
    from utils import load_cfg, load_weights, ensure_dir  # type: ignore

try:
    from model import build_model  # type: ignore
except Exception:
    try:
        from entry import build_model  # type: ignore
    except Exception:
        build_model = None


def build_model_from_cfg(cfg: Any) -> torch.nn.Module:
    choose_model = str(getattr(cfg.train, "choose_model", "Ours"))

    if "Ours" in choose_model:
        if build_model is None:
            raise ImportError("Cannot import build_model from your project.")
        model = build_model(cfg)
        print("[Model] Using Ours")
        return model

    if "SwinUNETR" in choose_model:
        from monai.networks.nets import SwinUNETR

        model = SwinUNETR(
            in_channels=len(cfg.data.use_modalities),
            out_channels=int(getattr(cfg.model, "out_ch", 1)),
            img_size=tuple(cfg.data.target_size),
            feature_size=96,
        )
        print("[Model] Using SwinUNETR")
        return model

    raise ValueError(f"Unsupported cfg.train.choose_model: {choose_model}")


@torch.no_grad()
def binary_dice_per_case(pred: torch.Tensor, target: torch.Tensor, eps: float = 1e-8) -> torch.Tensor:
    """
    pred/target: [B, 1, H, W, D] or compatible binary tensors.
    Returns: [B]
    """
    if pred.shape != target.shape:
        raise ValueError(f"Shape mismatch: pred={tuple(pred.shape)} target={tuple(target.shape)}")

    pred = pred.float().reshape(pred.shape[0], -1)
    target = target.float().reshape(target.shape[0], -1)

    inter = (pred * target).sum(dim=1)
    pred_sum = pred.sum(dim=1)
    target_sum = target.sum(dim=1)
    denom = pred_sum + target_sum

    dice = (2.0 * inter + eps) / (denom + eps)
    both_empty = (pred_sum == 0) & (target_sum == 0)
    dice = torch.where(both_empty, torch.ones_like(dice), dice)
    return dice


def _strip_nii_suffix(name: str) -> str:
    if name.endswith(".nii.gz"):
        return name[:-7]
    return Path(name).stem


def _get_meta(batch: Dict[str, Any], idx: int) -> Dict[str, Any]:
    meta = batch.get("meta", None)
    if isinstance(meta, list) and idx < len(meta) and isinstance(meta[idx], dict):
        return meta[idx]
    return {}


def safe_case_id_from_batch(batch: Dict[str, Any], idx: int) -> str:
    meta_i = _get_meta(batch, idx)
    case_id = meta_i.get("id", None)
    if case_id is not None:
        return str(case_id)

    raw_id = batch.get("id", None)
    if isinstance(raw_id, (list, tuple)) and idx < len(raw_id):
        return str(raw_id[idx])
    if raw_id is not None:
        return str(raw_id)

    return f"unknown_{idx}"


def safe_excel_id_from_batch(batch: Dict[str, Any], idx: int) -> str:
    meta_i = _get_meta(batch, idx)
    excel_id = meta_i.get("excel_id", None)
    if excel_id is not None:
        return str(excel_id)
    return ""


def safe_disk_case_name_from_batch(batch: Dict[str, Any], idx: int, modalities: Sequence[str]) -> str:
    meta_i = _get_meta(batch, idx)
    items = meta_i.get("items", {}) if isinstance(meta_i, dict) else {}

    for mod in modalities:
        item = items.get(f"image_{mod}", None)
        if isinstance(item, dict):
            filename = item.get("filename", None)
            if filename:
                return _strip_nii_suffix(Path(str(filename)).name)

    case_id = meta_i.get("id", None)
    if case_id is not None:
        return str(case_id)

    return safe_case_id_from_batch(batch, idx)


def _select_label_for_modality(y_all: torch.Tensor, mod_idx: int) -> torch.Tensor:
    if y_all.ndim != 5:
        raise ValueError(f"Expected label tensor 5D [B,C,H,W,D], got {tuple(y_all.shape)}")
    if y_all.shape[1] == 1:
        return y_all
    if mod_idx >= y_all.shape[1]:
        return y_all[:, 0:1]
    return y_all[:, mod_idx:mod_idx + 1]


def _select_pred_for_modality(pred_all: torch.Tensor, mod_idx: int) -> torch.Tensor:
    if pred_all.ndim != 5:
        raise ValueError(f"Expected pred tensor 5D [B,C,H,W,D], got {tuple(pred_all.shape)}")
    if pred_all.shape[1] == 1:
        return pred_all
    if mod_idx >= pred_all.shape[1]:
        return pred_all[:, 0:1]
    return pred_all[:, mod_idx:mod_idx + 1]


@torch.no_grad()
def evaluate_per_sample_dice(
    *,
    model: torch.nn.Module,
    val_loader,
    device: torch.device,
    threshold: float,
    modalities: Sequence[str],
) -> List[Dict[str, Any]]:
    model.eval()
    act = Activations(sigmoid=True)
    to_bin = AsDiscrete(threshold=float(threshold))

    rows: List[Dict[str, Any]] = []
    global_index = 0
    total_batches = len(val_loader) if hasattr(val_loader, "__len__") else None

    print("\n===== Start per-sample validation =====")
    if total_batches is not None:
        print(f"[Progress] total_batches={total_batches}")

    for batch_idx, batch in enumerate(val_loader):
        if batch is None or (not isinstance(batch, dict)):
            print(f"[Skip] batch_idx={batch_idx} invalid batch")
            continue
        if "image" not in batch or "seg_label" not in batch:
            print(f"[Skip] batch_idx={batch_idx} missing keys")
            continue

        x = batch["image"].to(device, non_blocking=True)
        y_all = batch["seg_label"].to(device, non_blocking=True)

        out = model(x)
        logits = out.logits if hasattr(out, "logits") else out
        prob = act(logits)
        pred_all = to_bin(prob)

        pred_sum_each = pred_all.reshape(pred_all.shape[0], -1).sum(dim=1)

        per_modality_dice: Dict[str, torch.Tensor] = {}
        per_modality_gt_sum: Dict[str, torch.Tensor] = {}

        for mod_idx, mod in enumerate(modalities):
            pred_mod = _select_pred_for_modality(pred_all, mod_idx)
            y_mod = _select_label_for_modality(y_all, mod_idx)
            per_modality_dice[mod] = binary_dice_per_case(pred_mod, y_mod)
            per_modality_gt_sum[mod] = y_mod.reshape(y_mod.shape[0], -1).sum(dim=1)

        for i in range(x.shape[0]):
            disk_case_name = safe_disk_case_name_from_batch(batch, i, modalities)
            case_id = safe_case_id_from_batch(batch, i)
            excel_id = safe_excel_id_from_batch(batch, i)
            pred_voxels = float(pred_sum_each[i].item())

            row: Dict[str, Any] = {
                "sample_index": int(global_index),
                "batch_idx": int(batch_idx),
                "sample_idx_in_batch": int(i),
                "disk_case_name": str(disk_case_name),
                "case_id": str(case_id),
                "excel_id": str(excel_id),
                "pred_voxels": pred_voxels,
                "is_pred_empty": bool(pred_voxels == 0),
            }

            dice_values: List[float] = []
            msg_parts: List[str] = []

            for mod in modalities:
                dice_value = float(per_modality_dice[mod][i].item())
                gt_voxels = float(per_modality_gt_sum[mod][i].item())
                row[f"dice_{mod}"] = dice_value
                row[f"gt_voxels_{mod}"] = gt_voxels
                row[f"is_gt_empty_{mod}"] = bool(gt_voxels == 0)
                dice_values.append(dice_value)
                msg_parts.append(f"dice_{mod}={dice_value:.6f}")

            mean_dice = float(sum(dice_values) / max(len(dice_values), 1))
            row["dice_mean"] = mean_dice
            row["num_modalities"] = int(len(modalities))
            rows.append(row)

            progress_prefix = f"[Val {batch_idx + 1}/{total_batches}]" if total_batches is not None else "[Val]"
            print(
                f"{progress_prefix} sample_index={global_index} | "
                f"disk_case_name={disk_case_name} | excel_id={excel_id} | "
                + " | ".join(msg_parts)
                + f" | dice_mean={mean_dice:.6f} | pred_voxels={pred_voxels:.0f}"
            )

            global_index += 1

    rows.sort(key=lambda x: (x["dice_mean"], x["sample_index"]))
    return rows


def _get_nested(cfg: Any, path: str, default: Any = None) -> Any:
    cur = cfg
    for key in path.split("."):
        if cur is None:
            return default
        if isinstance(cur, dict):
            cur = cur.get(key, None)
        else:
            cur = getattr(cur, key, None)
        if cur is None:
            return default
    return cur


def resolve_output_paths(cfg: Any, cli_out_dir: str, cli_worst_k: int) -> Dict[str, Any]:
    cfg_out_dir = _get_nested(cfg, "eval.per_sample_dice.out_dir", "")
    cfg_xlsx_name = _get_nested(cfg, "eval.per_sample_dice.xlsx_name", "val_per_sample_dice.xlsx")
    cfg_csv_name = _get_nested(cfg, "eval.per_sample_dice.csv_name", "val_per_sample_dice.csv")
    cfg_json_name = _get_nested(cfg, "eval.per_sample_dice.json_name", "val_per_sample_dice.json")
    cfg_txt_name = _get_nested(cfg, "eval.per_sample_dice.summary_name", "val_per_sample_dice_summary.txt")
    cfg_low_xlsx_name = _get_nested(cfg, "eval.per_sample_dice.low_dice_xlsx_name", "val_per_sample_low_dice.xlsx")
    cfg_low_csv_name = _get_nested(cfg, "eval.per_sample_dice.low_dice_csv_name", "val_per_sample_low_dice.csv")
    cfg_low_json_name = _get_nested(cfg, "eval.per_sample_dice.low_dice_json_name", "val_per_sample_low_dice.json")
    cfg_worst_k = int(_get_nested(cfg, "eval.per_sample_dice.worst_k", cli_worst_k))
    cfg_low_dice = float(_get_nested(cfg, "eval.per_sample_dice.low_dice", 0.60))

    out_dir = Path(str(cli_out_dir).strip() or str(cfg_out_dir).strip() or "./analysis_per_sample_dice")
    out_dir = ensure_dir(out_dir)

    return {
        "out_dir": out_dir,
        "xlsx_path": out_dir / str(cfg_xlsx_name),
        "csv_path": out_dir / str(cfg_csv_name),
        "json_path": out_dir / str(cfg_json_name),
        "txt_path": out_dir / str(cfg_txt_name),
        "low_xlsx_path": out_dir / str(cfg_low_xlsx_name),
        "low_csv_path": out_dir / str(cfg_low_csv_name),
        "low_json_path": out_dir / str(cfg_low_json_name),
        "worst_k": cfg_worst_k,
        "low_dice": cfg_low_dice,
    }


def _write_excel_as_text(df: pd.DataFrame, xlsx_path: Path) -> None:
    df_to_write = df.copy()
    for col in df_to_write.columns:
        if col in {"disk_case_name", "case_id", "excel_id"}:
            df_to_write[col] = df_to_write[col].fillna("").map(str)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_to_write.to_excel(writer, index=False, sheet_name="dice")

    wb = load_workbook(xlsx_path)
    ws = wb["dice"]
    text_cols = {"disk_case_name", "case_id", "excel_id"}
    header_to_col = {ws.cell(row=1, column=j).value: j for j in range(1, ws.max_column + 1)}

    for name in text_cols:
        col_idx = header_to_col.get(name, None)
        if col_idx is None:
            continue
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = "@"
            if cell.value is None:
                cell.value = ""
            else:
                cell.value = str(cell.value)

    wb.save(xlsx_path)


def save_reports(
    rows: List[Dict[str, Any]],
    *,
    xlsx_path: Path,
    csv_path: Path,
    json_path: Path,
    txt_path: Path,
    low_xlsx_path: Path,
    low_csv_path: Path,
    low_json_path: Path,
    worst_k: int,
    low_dice: float,
    modalities: Sequence[str],
) -> Dict[str, Path]:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    columns = [
        "sample_index",
        "batch_idx",
        "sample_idx_in_batch",
        "disk_case_name",
        "case_id",
        "excel_id",
    ]
    for mod in modalities:
        columns.extend([f"dice_{mod}", f"gt_voxels_{mod}", f"is_gt_empty_{mod}"])
    columns.extend([
        "dice_mean",
        "num_modalities",
        "pred_voxels",
        "is_pred_empty",
    ])

    df = pd.DataFrame(rows, columns=columns)
    low_df = df[df["dice_mean"] < float(low_dice)].copy()

    _write_excel_as_text(df, xlsx_path)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    _write_excel_as_text(low_df, low_xlsx_path)
    low_df.to_csv(low_csv_path, index=False, encoding="utf-8-sig")
    low_json_rows = low_df.to_dict(orient="records")
    low_json_path.write_text(json.dumps(low_json_rows, ensure_ascii=False, indent=2), encoding="utf-8")

    mean_dice = float(df["dice_mean"].mean()) if len(df) > 0 else 0.0
    worst_n = min(int(worst_k), len(df))

    lines = [
        "Per-sample Dice Summary",
        f"total_cases: {len(df)}",
        f"mean_dice: {mean_dice:.6f}",
        f"low_dice_threshold: {float(low_dice):.6f}",
        f"low_dice_count: {len(low_df)}",
        f"worst_k: {int(worst_k)}",
        "",
        f"Worst {worst_n} by dice_mean:",
    ]

    for i, (_, r) in enumerate(df.head(worst_n).iterrows(), 1):
        parts = [
            f"{i:02d}. sample_index={int(r['sample_index'])}",
            f"disk_case_name={r['disk_case_name']}",
            f"excel_id={r['excel_id']}",
            f"dice_mean={float(r['dice_mean']):.6f}",
        ]
        for mod in modalities:
            parts.append(f"dice_{mod}={float(r[f'dice_{mod}']):.6f}")
        lines.append(" | ".join(parts))

    txt_path.write_text("\n".join(lines), encoding="utf-8")

    return {
        "xlsx": xlsx_path,
        "csv": csv_path,
        "json": json_path,
        "summary": txt_path,
        "low_dice_xlsx": low_xlsx_path,
        "low_dice_csv": low_csv_path,
        "low_dice_json": low_json_path,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compute Dice for each validation sample.")
    parser.add_argument("--config", type=str, default="config.yml")
    parser.add_argument("--weights", type=str, default="", help="Override cfg.eval.weights_path")
    parser.add_argument("--device", type=str, default="cuda")
    parser.add_argument("--threshold", type=float, default=None, help="Override cfg.eval.threshold")
    parser.add_argument("--out-dir", type=str, default="", help="Override cfg.eval.per_sample_dice.out_dir")
    parser.add_argument("--worst-k", type=int, default=20)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    cfg = load_cfg(args.config)

    weights_path = str(args.weights).strip() or str(getattr(cfg.eval, "weights_path", "")).strip()
    if not weights_path:
        raise ValueError("No weights path provided. Set --weights or cfg.eval.weights_path.")

    threshold = float(args.threshold) if args.threshold is not None else float(getattr(cfg.eval, "threshold", 0.5))

    if str(args.device).lower() == "cuda" and not torch.cuda.is_available():
        print("[Warn] CUDA not available, fallback to CPU.")
        device = torch.device("cpu")
    else:
        device = torch.device(args.device)

    out_cfg = resolve_output_paths(cfg, cli_out_dir=args.out_dir, cli_worst_k=int(args.worst_k))
    modalities = [str(x) for x in getattr(cfg.data, "use_modalities", [])]
    if len(modalities) == 0:
        raise ValueError("cfg.data.use_modalities is empty.")

    print(f"[Config] {args.config}")
    print(f"[Weights] {weights_path}")
    print(f"[Device] {device}")
    print(f"[Threshold] {threshold}")
    print(f"[Modalities] {modalities}")
    print(f"[Output Dir] {out_cfg['out_dir']}")
    print(f"[Low Dice Threshold] {float(out_cfg['low_dice']):.6f}")

    if cfg.eval.train != True:
        _, val_loader = get_loaders(cfg.data)
    else:
        val_loader, _ = get_loaders(cfg.data)

    model = build_model_from_cfg(cfg)
    info = load_weights(model=model, weights_path=weights_path, strict=False, map_location="cpu")
    print(f"[Load] missing={len(info['missing_keys'])} | unexpected={len(info['unexpected_keys'])}")

    model.to(device)

    rows = evaluate_per_sample_dice(
        model=model,
        val_loader=val_loader,
        device=device,
        threshold=threshold,
        modalities=modalities,
    )

    if len(rows) == 0:
        raise RuntimeError("No validation samples were evaluated.")

    paths = save_reports(
        rows,
        xlsx_path=out_cfg["xlsx_path"],
        csv_path=out_cfg["csv_path"],
        json_path=out_cfg["json_path"],
        txt_path=out_cfg["txt_path"],
        low_xlsx_path=out_cfg["low_xlsx_path"],
        low_csv_path=out_cfg["low_csv_path"],
        low_json_path=out_cfg["low_json_path"],
        worst_k=int(out_cfg["worst_k"]),
        low_dice=float(out_cfg["low_dice"]),
        modalities=modalities,
    )

    mean_dice = sum(r["dice_mean"] for r in rows) / len(rows)
    low_rows = [r for r in rows if r["dice_mean"] < float(out_cfg["low_dice"])]

    print("\n===== Per-sample Dice Summary =====")
    print(f"Total cases: {len(rows)}")
    print(f"Mean dice_mean: {mean_dice:.6f}")
    print(f"Low dice threshold: {float(out_cfg['low_dice']):.6f}")
    print(f"Low dice cases: {len(low_rows)}")
    print(f"Worst {min(int(out_cfg['worst_k']), len(rows))} cases by dice_mean:")
    for i, r in enumerate(rows[: min(int(out_cfg["worst_k"]), len(rows))], 1):
        modal_bits = " | ".join([f"dice_{m}={r[f'dice_{m}']:.6f}" for m in modalities])
        print(
            f"{i:02d}. sample_index={r['sample_index']} | disk_case_name={r['disk_case_name']} | "
            f"excel_id={r['excel_id']} | {modal_bits} | dice_mean={r['dice_mean']:.6f}"
        )

    print("\nSaved files:")
    for k, p in paths.items():
        print(f"- {k}: {p}")


if __name__ == "__main__":
    main()
